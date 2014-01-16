#! /usr/bin/env python3.3
# -*- coding: UTF-8 -*-

from openpyxl import load_workbook
#from datetime import datetime

class TableReader:
	"""read table data from excel sheet"""
	def __init__(self, sheet,col_begin,col_end):
		self.sheet = sheet
		self.col_begin = col_begin
		self.col_end = col_end
		self.header = []

		self._build_header()

	def read(self):
		#max_row_index = self.sheet.get_highest_row()

		datas=[]
		ri=1
		for row in self.sheet.iter_rows():
			entity={}
			ci=0
			for cell in row:#range(self.col_begin,self.col_end):
				if(ci>=self.col_begin and ci<=self.col_end):
				#	cell = self.sheet.cell(row=ri,column=ci)
					# str_value=""
					# if cell.is_date():
					# 	str_value=cell.internal_value.isoformat()
					# elif cell.internal_value is None:
					# 	str_value=''
					# else:
					# 	str_value=str(cell.internal_value)

					entity[self.header[ci-self.col_begin]] = self._escape_sql_char(cell.internal_value)
				ci+=1
			
			datas.append(entity)
			ri+=1		

		return datas

	def _build_header(self):
		# for hi in range(self.col_begin,self.col_end):
		# 	header_cell = self.sheet.cell(row=0,column=hi)
		# 	self.header.append(header_cell.value)
		# hi=0
		for row in self.sheet.iter_rows():
			ci=0
			for cell in row:
				if (ci>=self.col_begin and ci<=self.col_end):
					self.header.append(cell.internal_value)
				ci+=1
			break		

	def get_header(self):
		return self.header


	def _escape_sql_char(self,sql):
		sql = str(sql)
		sql = sql.replace("'","\\'")
		sql = sql.replace('"','\\"')
		#sql = str(sql.encode('utf-8'))

		return sql


		